import graphs, GLSLio, FFio, analysis, GLSLutils, HYDATio, ECio
from matplotlib import pyplot as plt
import numpy as np
import pandas as pd
import os, pickle, json
from imp import reload

reload(graphs)
reload(analysis)
reload(GLSLio)


def station_level_stats():
	for key, sid in GLSLio.stations_niveaux.items():
		fig, ax = graphs.plot_station_qom_level_stats(sid)
		fig.savefig('../rapport/figs/stats_niveaux_{0}.png'.format(key))
		plt.close()
#
def scenario_depths():
	for scen in range(1,9):
		fig, ax = graphs.plot_depth_map(scen)
		fig.savefig('../rapport/figs/depth_scen_{0}.png'.format(scen), dpi=200)
		plt.close()
#
def FF_scenarios():
	for site in ['ont']:
		fig, ax = graphs.plot_FF_flow(site)
		fig.savefig('../rapport/figs/FF_scenarios_{0}.png'.format(site), dpi=200)
		plt.close()
#
def plot_mesh():
	fig, ax = graphs.plot_mesh()
	fig.savefig('../rapport/figs/mesh.png', dpi=200)

def plot_Sorel():
	fig, axes = graphs.plot_Sorel_annual_minimum_qom_levels_flows()
	fig.savefig('../rapport/figs/Sorel_flow_levels.png', dpi=200)
	plt.close()
#
def plot_stations_municipales():
	sites = GLSLio.stations_potables()
	pts = {}
	for key, val in sites.items():
		try:
			lon, lat = val[1]
			pts[key]= (lon, lat)
		except:
			pass
	pts.update( dict(zip(analysis.CP['sites'], analysis.CP['coords'])) )

	fig, ax = graphs.plot_depth_map(8, pts=pts, inset=False)
	return fig
#
def plot_stations_usees():
	sites = GLSLio.stations_usees_2()
	pts = {}
	for key, val in sites.items():
		lat, lon = val['lat'], val['lon']
		if lat is None or lon is None:
			continue
		pts[key] = (lon, lat)

	fig, ax = graphs.plot_depth_map(8, pts=pts, inset=False)
	return fig
#
## Senario #1 for Pointe-Claire

def AECOM_potable():

	meta = GLSLio.stations_potables()

	# Interpolate the levels
	ECLpath = '../analysis/potableECL_N.pickle'
	if os.path.exists(ECLpath):
		ECL = pickle.load(open(ECLpath, 'br'))
	else:
		compute_ECL_potable()
		add_ECL_carrieres()
		ECL = pickle.load(open(ECLpath, 'br'))

	out = {'REF':{} ,'WI1':{}, 'WI2':{}}


#	ks = list(ECL.keys())
#	for key in ks:
#		if key not in range(7, 50):
#			ECL.pop(key)

	for (key, Ls) in ECL.items():
		print(key)
		for i in [1,2,3]:
			L = Ls.get(i)
			if L is not None:
				lon, lat = meta[key][i]
				out['REF']['{0}_NO{1}'.format(key, i)] = analysis.scenario_HR(L)
				out['WI1']['{0}_NO{1}'.format(key, i)] = analysis.scenario_H1(None, lon, lat, L, key==17)
				out['WI2']['{0}_NO{1}'.format(key, i)] = analysis.scenario_H2(L)


	for key in out['REF'].keys():
		fn = '../deliverables/Eaux/Potable/Scenarios_niveaux_eau_potable_{0}_IGLD85.xlsx'.format(key)
		EW = pd.ExcelWriter(fn)
		for i, scen in enumerate(['REF', 'WI1', 'WI2']):
			out[scen][key].to_frame('Level m').to_excel(EW, scen)
		EW.close()

	return out

def AECOM_usee():
	scens = ['REF', 'WI1', 'WI2']
	meta = GLSLio.stations_usees_2()

	contributions = {
	(8,12,7,26): ["LaSalle",],
	(18,27,22,9,6,5): ["LaSalle", "MIP", "Assomption"],
	(20,25): ["LaSalle", "MIP", "Assomption", "Richelieu"],
	(1,): ["LaSalle", "MIP", "Assomption", "Richelieu", "Yamaska", "St-Francois", "Maskinongé", "duLoup"],
	(14,17):["LaSalle", "MIP", "Assomption", "Richelieu", "Yamaska", "St-Francois", "Maskinongé", "duLoup", "Nicolet"],
	(11,16): ["Ch. Ste-Anne-Vaudreuil",],
	(13,): ["lesCedres",],
	(2,): ["Beauharnois",],
	(23,) : ["lesCedres", "Beauharnois"]
	}

	ECQ = {}
	for key in meta.keys():
		for k,v in contributions.items():
			if key in k:
				ECQ[key] = sum([np.array(analysis.EC_scen_Q[site]) for site in v])
				break

	# Interpolate the levels
	ECLpath = '../analysis/useesECL_N.pickle'
	if os.path.exists(ECLpath):
		ECL = pickle.load(open(ECLpath, 'br'))
	else:
		ECL = {}
		for k, v in meta.items():
			lat, lon = v['lat'], v['lon']
			try:
				ECL[k] = analysis.interpolate_EC_levels(lon, lat).tolist()
			except ValueError:
				ECL[k] = None
		pickle.dump(ECL, open(ECLpath, 'bw'))

	out = {}
	for (key, val) in meta.items():
		lon, lat = val['lon'], val['lat']
		L = ECL.get(key)
		Q = ECQ.get(key)
		print(key)
		if L is not None:
			out[key] = analysis.scenarios_QH(None, lon, lat, Q, L)

	for key, res in out.items():
		fn = '../deliverables/Eaux/Usees/Scenarios_niveaux-debits_eau_usees_{0}_IGLD85.xlsx'.format(key)
		EW = pd.ExcelWriter(fn)

		for i, scen in enumerate(['REF', 'WI1', 'WI2']):
			res[i].to_excel(EW, scen)

		EW.close()

	return out



def compute_ECL_potable():
	ECL = {}
	meta = GLSLio.stations_potables()
	for k, v in meta.items():
		ECL[k] = {}
		for i in [1,2,3]:
			if meta[k].get(i) is None:
				continue

			lon, lat = meta[k][i]
			try:
				ECL[k][i] = analysis.interpolate_EC_levels(lon, lat).tolist()
			except ValueError:
				pass

	pickle.dump(ECL, open('../analysis/potableECL_N.pickle', 'bw'))

def write_domain_in_usees():
	from openpyxl import Workbook, load_workbook
	filename='../data/AECOM/Calculs wwtp database check.xlsx'
	wb = load_workbook(filename=filename)
	ws = wb['wwtp']
	sites = analysis.check_usines_usees_in_domain()

	ws['Q1'] = 'Domain'
	for key, val in sites.items():
		ws.cell(row=key+1, column=17).value = val['dom']

	wb.save('../data/AECOM/Calculs wwtp database domain check.xlsx')

def write_domain_in_potables():
	from openpyxl import Workbook, load_workbook

	sites = GLSLio.stations_potables()
	L = pickle.load(open('../analysis/potableECL_N.pickle'))

	filename='../data/AECOM/Calculs UTEP.xlsx'
	wb = load_workbook(filename=filename)
	ws = wb['UTEP']

	cols = ['', 'AV', 'AW', 'AX']

	for i in range(1,4):
		ws['{0}{1}'.format(cols[i],1)] = "Domain{0}".format(i)
		for key, v in sites.items():
			if L[key].get(i) is not None:
				lon, lat = L[key][i]
				ws['{0}{1}'.format(cols[i], key+1)] = analysis.get_domain(lon, lat)

	wb.save('../data/AECOM/Calculs UTEP domain check.xlsx')


def add_ECL_carrieres():
	ECLpath = '../analysis/potableECL_N.pickle'
	ECL = pickle.load(open(ECLpath, 'br'))

	# 12 : La Prairie
	ECL[12][1] = [8.3972052587, 9.0669104397, 9.6616692046, 10.2029027485, 11.0188768969, 11.7560161993, 12.5646368075, 13.3103094861]

	# 18 : Montréal Atwater
	ECL[18][1] = [17.5880882, 18.0475934, 18.555371, 18.990474, 19.633782, 20.1145572, 20.7504656, 21.2578254]

	# 6: Candiac
	ECL[6][1] = [8.7882906836, 9.4553778364, 10.046259892, 10.5828876396, 11.3902671625, 12.1181908673, 12.915365757, 13.6494202691]

	pickle.dump(ECL, open(ECLpath, 'bw'))


def Laura():
	import ECio
	bc_EC_Q = ECio.Q_Sorel('qtm')
	ECL = [ 3.09323328,  3.49145638,  4.1352682 ,  4.74138233,  5.49859576,
	6.38552684,  7.10856182,  8.12003153]

	bc_EC = pd.Series(analysis.weight_EC_levels(ECL, analysis.get_EC_scenario_index(bc_EC_Q)), bc_EC_Q.index)

	fr = bc_EC.to_frame('Niveau' + ' m')
	fr['Debit m3s'] = bc_EC_Q
	fr.to_excel('../deliverables/Sorel_Laura.xlsx')



def dump_xls(flow, level, filename, offset=0):
	Fr = level.valid().to_frame('Level m')
	Fr['Flow m3s'] = flow
	y, q = Fr.index.levels
	Fr.index.set_levels((y+offset, q), True)
	out = Fr

	Fr = Fr.swaplevel(0,1)

	P = Fr.to_panel()
	P.to_excel(filename)
	return out


def ECOSYSTEMES():
	# Datum ID: 431, Great Lakes International Datum 1985
	# From EC datum reference:
	# Station	Nom de la station	RN de référence	Altitude(ZC)	Altitude(SRIGL85)	SRIGL85(ZC)
	# 15975	  Lac Saint-Pierre	 87L9000			6.537		   9.980				3.443

	# Les données du modèles sont en Niveau Moyen des Mers, ce qui je pense,
	# correspond à l'altitude. Si on veut reporter ces valeurs au datum de la
	# station de mesure, on doit donc soustraire 6.537 des scénarios.

	# Level observations at station
	# 02OC016	SAINT-PIERRE (LAC) AU CURB NO. 2
	# Stations H Lac St-Pierre

	s = '02OC016'
	M = HYDATio.get_station_meta(s)
	lat, lon = M['LATITUDE'], M['LONGITUDE']
	ECH = analysis.interpolate_EC_levels(lon, lat)
	ECQ = analysis.EC_scen_Q['Trois-Rivieres']

	res = analysis.scenarios_QH(site='lsp', ECQ=ECQ, ECH=ECH )

	fn = '../deliverables/Ecosystemes/Scenarios_niveaux-debits_Lac_St-Pierre_02OC016_IGLD85.xlsx'
	EW = pd.ExcelWriter(fn)
	for i, scen in enumerate(['REF', 'WI1', 'WI2']):
		res[i].to_excel(EW, scen)

	EW.close()

	pickle.dump(res, open('../analysis/ecosystemes.pickle', 'wb'))
	return res

#
def TOURISME():
	"""Certains des positions des marinas correspondent aux batiments, et non au bassin de la marina.
	J'ai contacté Stéphanie pour qu'on décide ce qu'on fait avec ca."""
	meta = GLSLio.marinas()

	# Interpolate the levels at the marinas
	ECLpath = '../analysis/marinasECL_N.json'
	if os.path.exists(ECLpath):
		ECL = json.load(open(ECLpath))
	else:

		ECL = {}
		for k, v in meta.items():
			try:
				ECL[k] = analysis.interpolate_EC_levels(*v).tolist()
			except ValueError:
				ECL[k] = None
		json.dump(ECL, open(ECLpath, 'w'))

	out = {'REF':{} ,'WI1':{}, 'WI2':{}}

	for (key, L) in ECL.items():
		if L is not None:
			print(key)
			lon, lat = meta[key]
			out['REF'][key] = analysis.scenario_HR(L)
			out['WI1'][key] = analysis.scenario_H1(None, lon, lat, L)
			out['WI2'][key] = analysis.scenario_H2(L)

	for (key, L) in ECL.items():
		if L is not None:
			fn = '../deliverables/Tourisme/Scenarios_niveaux_{0}_IGLD85.xlsx'.format(key)
			EW = pd.ExcelWriter(fn)
			for i, scen in enumerate(['REF', 'WI1', 'WI2']):
				out[scen][key].to_frame('Level m').to_excel(EW, scen)
			EW.close()

	return out

def TOURISME2():
	"""
	1. pour les 4 endroits suivants (peu importe où dans la zone)
	a. Lac St-François
	b. Lac St-Louis
	c. Montréal-Boucherville-Contrecoeur
	d. Sorel-Lac St-Pierre

	2. Sur la période retenue 2015-2065
	Le nombre de jours par année/mois sur la période de navigation de plaisance avril-novembre où l'on peut s'attendre à 
	a) -10cm
	b) -20cm à 29cm
	c) -30cm à -39cm
	d) > que -40cm

	"""
	from collections import OrderedDict

	args = [('02OA039', 20.351, 'Pointe-Claire'),
			('02OC016', 3.443, 'Lac St-Pierre'),
			('02OJ033', 3.8, 'Sorel')]

	for arg in args:
		s, zc, name = arg
		meta = HYDATio.get_station_meta(s)

		lon, lat = meta['LONGITUDE'], meta['LATITUDE']
		EC = analysis.interpolate_EC_levels(lon, lat).tolist()
		scens = analysis.scenarios_H(lat=lat, lon=lon, EC=EC)

		fn = '../deliverables/Tourisme/Scenarios_niveaux_jours_{0}.xlsx'.format(name)
		EW = pd.ExcelWriter(fn)

		for scen, s in zip(("REF", "WI1", "WI2"), scens):
			s -= zc

			data = OrderedDict()
			data['<-40'] = s < -.40
			data['[-40, -30['] = (s>=-.40) & (s<-.30)
			data['[-30, -20]'] = (s>=-.30) & (s<-.20)
			data['[-20, -10]'] = (s>=-.20) & (s<-.10)
			data['[-10, 0]'] = (s>=-.10) & (s<0)
			data['[0, 10]'] = (s>=0) & (s<.10)

			df = pd.DataFrame(data=data)
			ndays = GLSLutils.qomdays(df.index)
			df = df.apply(lambda x: x * ndays)
			df.reset_index(inplace=True)
			df['Month'] = (df['QTM']-1)//4 + 1
			df.drop('QTM', axis=1, inplace=True)
			df = df.query('Month >=4 & Month <=11')
			out = df.groupby(['Year', 'Month'], as_index=True).sum()

			out.to_excel(EW, scen)

		EW.close()




def TRANSPORT():
	"""Niveaux a la jetee #1.

	Relation 0 des cartes IGLD85: 5.560
	"""
	zc = 5.56
	ECL = analysis.EC_scen_L['mtl']

	out = {}

	out['REF'], out['WI1'], out['WI2'] = analysis.scenarios_H('mtl', EC=ECL)

	# Series
	fn = '../deliverables/Transport/Scenarios_series_niveaux_Jetee_1_mtl.xlsx'
	EW = pd.ExcelWriter(fn)

	for i, scen in enumerate(['REF', 'WI1', 'WI2']):
		out[scen] = out[scen] - zc
		out[scen].to_frame('Level m').to_excel(EW, scen)
	EW.close()

	# Panels
	fn = '../deliverables/Transport/Scenarios_tableaux_niveaux_Jetee_1_mtl.xlsx'
	EW = pd.ExcelWriter(fn)


	for i, scen in enumerate(['REF', 'WI1', 'WI2']):
		F = out[scen].to_frame(scen)
		y, q = F.index.levels
		F.index.set_levels((y, q), True)
		F = F.swaplevel(0,1)

		F.to_panel().to_excel(EW, scen)
		EW.close()

	return out


def FONCIER():
	"""Niveaux au Lac St-Louis (Pointe-Claire)"""
	fn = '../deliverables/Foncier/Scenarios_niveaux_Pointe-Claire_02OA039.xlsx'
	EW = pd.ExcelWriter(fn)
	out = {}

	# Observations at Pointe-Claire
	s = '02OA039'

	M = HYDATio.get_station_meta(s)
	lat, lon = M['LATITUDE'], M['LONGITUDE']
	ECL = analysis.interpolate_EC_levels(lon, lat)

	if 0: #OBS
		L = HYDATio.get_hydat(s, 'H')
		LQ = GLSLutils.group_qom(L).mean()
		LQ.index.names = ["Year", "QTM"]
		F = LQ.to_frame('Level m')
		out['obs'] = F

	EW = pd.ExcelWriter(fn)

	scens = analysis.scenarios_H(site='pcl', EC=ECL)
	k = "Level m"
	for scen, ts in zip(["REF", "WI1", "WI2"], scens):
		ts.to_frame(k).to_excel(EW, scen)

	#	qbc.to_frame("REF + m3s").to_excel(EW, 'REF')#
	#wd.to_frame("WI1 + m3s").to_excel(EW, 'WI1')
	#q2.to_frame("WI2 + m3s").to_excel(EW, 'WI2')

	EW.close()


	return out

def FONCIER2():
	from osgeo import ogr
	import fiona
	from matplotlib import pyplot as plt
	import shapely

	for reg in ['lsl', 'mtl_lano', 'lsp']:

		T = analysis.get_tesselation(reg)

		features = []
		for i in range(8):
			d = ECio.EC_depth(reg, i+1)
			C = plt.tricontour(T, d, [0,])
			lc = C.collections[0] # line collection
			seg = lc.get_segments()
			plt.close()

			features.append( {'geometry': {'type': 'MultiLineString',
							 'coordinates': seg},
							'properties': { 'SCEN_ID': str(i+1), \
											'FLOW':analysis.EC_scen_Q['Sorel'][i]}})


		with fiona.open('../deliverables/Foncier/GLSL_contour_{0}.tab'.format(reg), 'w',\
						crs={'no_defs': True, 'init': 'epsg:32188'},\
						driver="MapInfo File",\
						schema={'geometry':'MultiLineString',\
							   'properties':{'SCEN_ID': 'str',\
											 'FLOW': 'float:16'}}) as sink:

			sink.writerecords(features)

def FONCIER2xls():
	from matplotlib import pyplot as plt
	import shapely
	from openpyxl import Workbook

	wb = Workbook()
	reg = 'lsl'
	T = analysis.get_tesselation(reg, option='B')

	features = []
	for i in range(8):
		ws = wb.create_sheet(index=i, title='S'+str(i+1))
		d = ECio.EC_depth(reg, i+1)
		C = plt.tricontour(T, d, [0,])
		lc = C.collections[0] # line collection
		seg = np.vstack( lc.get_segments() )
		plt.close()
		ws.append(['X', 'Y'])
		for row in seg.tolist():
			ws.append(row)


	wb.save('../deliverables/Foncier/GLSL_contour.xlsx')

def FONCIER3():

	fn = '../deliverables/Foncier/Scenarios_debits_Sorel.xlsx'
	EW = pd.ExcelWriter(fn)

	scens = analysis.scenarios_Q(site='srl')
	k = "Flow m3s"
	for scen, ts in zip(["REF", "WI1", "WI2"], scens):
		ts.to_frame(k).to_excel(EW, scen)

	EW.close()

def FONCIER4():

	fn = '../deliverables/Foncier/Scenarios_debits_Sorel.xlsx'
	EW = pd.ExcelWriter(fn)

	scens = analysis.scenarios_Q(site='srl')
	k = "Flow m3s"
	for scen, ts in zip(["REF", "WI1", "WI2"], scens):
		#ts.groupby(lambda x: (x[0], np.floor((x[1]-1)/4))).mean()
		mm = pd.rolling_mean(ts, 4)[3::4]
		mi = pd.MultiIndex.from_tuples([(y,int(m/4)) for (y,m) in mm.index.values], names=('Year', 'Month'))
		mts = pd.TimeSeries(mm.values, index=mi)

		df = mts.to_frame(k)
		wi = analysis.get_EC_scenario_index(mts)
		df['Poids'] = pd.TimeSeries(wi, index=mi)
		df['Index'] = pd.TimeSeries(np.around(wi), index=mi)
		df.to_excel(EW, scen)

	EW.close()

def HYDRO():
	"""Scénario de débit pour Beauharnois et Les Cedres."""


	fn = '../deliverables/Hydroelectricite/Scenarios_debit_lac_Ontario.xlsx'
	EW = pd.ExcelWriter(fn)
	tmp = {}

	ECQ = np.array( analysis.EC_scen_Q['Beauharnois'] ) + np.array( analysis.EC_scen_Q['lesCedres'] )

	res = analysis.scenarios_Q(EC=ECQ)

	for i, scen in enumerate(['REF', 'WI1', 'WI2']):
		res[i].to_frame('Flow m3s').to_excel(EW, scen)
	EW.close()
	return res
#
def plot_ww_wd_scenarios():
	fig, ax = graphs.plot_FF_flow('ont')
	fig.savefig('../figs/FF_wd_ww_scenarios_ontario.png')
	plt.close()

#
def plot_aecom():
	from matplotlib.transforms import offset_copy
	data = pickle.load(open('../analysis/aecom.pickle', 'rb'))
	fig = graphs.scenarios(data)

	fig.axes[0].text(0, 1, "Pointe-Claire", fontsize=34, weight='bold', color='#2b2929', alpha=.8, transform=offset_copy(fig.axes[0].transAxes, x=0, y=30, units='dots'))
	fig.savefig('../figs/Pointe-Claire.png')
#
def plot_ecosystemes():
	from matplotlib.transforms import offset_copy
	data = pickle.load(open('../analysis/ecosystemes.pickle', 'rb'))
	fig = graphs.scenarios(data)

	fig.axes[0].text(0, 1, "Lac St-Pierre", fontsize=34, weight='bold', color='#2b2929', alpha=.8, transform=offset_copy(fig.axes[0].transAxes, x=0, y=30, units='dots'))
	fig.savefig('../figs/Lac St-Pierre.png')
#
def ff_xls(site, save=False):
	from pandas.io.excel import ExcelWriter

	CP = analysis.CP
	i = CP['sites'].index(site)
	#W = ExcelWriter(fn)

	# Corresponding flow
	if site in ['var', 'pcl']:
		fs = ('stl',)
	elif site == 'srl':
		fs = ('stl', 'dpmi')


	out = {}
	for scen in ['bc', 'wd']:
		# Get level
		ts = FFio.level_series_QH(site, scen)

		# Sum flow contributions
		f = FFio.FF_flow(fs[0], scen)
		for s in fs[1:]:
			f = f.add( FFio.FF_flow(s, scen) )

		F = ts.valid().to_frame('Level m')
		F['Flow m3s'] = f
		y, q = F.index.levels
		F.index.set_levels((y+analysis.offset[scen], q), True)
		F = F.swaplevel(0,1)

		if save:
			P = F.to_panel()
			fn = 'Water_Levels_IJC_{0}_{1}_IGLD85.xlsx'.format(CP['names'][i], scen.upper())
			P.to_excel(fn)

		out[scen] = F
	if not save:
		return out


def chee():
	for site in ['srl', 'var']:
		ff_xls(site)
