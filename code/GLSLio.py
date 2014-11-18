"""
Lecture des données de niveaux et débits d'EC.

David Huard, 2014

"""

"""Notes sur la base de données Hydat.

Téléchargement: 25 mars 2014.
Conversion sqlite: AccessDump.py
Commande SQL pour identifier les stations d'intérêt pour les débits:

SELECT STATION_NUMBER, STATION_NAME from "stations" WHERE STATION_NUMBER IN (
SELECT  STATION_NUMBER FROM "stations" WHERE "STATION_NAME" LIKE "%FLEUVE%"
INTERSECT
SELECT STATION_NUMBER FROM "stn_data_range" WHERE RECORD_LENGTH>30  AND DATA_TYPE='Q');
#======#
02MC010	SAINT-LAURENT (FLEUVE) - CENTRALE DE BEAUHARNOIS
02MC018	SAINT-LAURENT (FLEUVE) - CENTRALE DES CEDRES
02MC024	SAINT-LAURENT (FLEUVE)(CHENAL BEAUHARNOIS) LAC SAINT FRANCOIS
02MC029	SAINT-LAURENT (FLEUVE) AU BARRAGE DE L'ILE
02OA016	SAINT-LAURENT (FLEUVE) A LASALLE
02OC003    SAINTE-PIERRE (LAC) A LOUISEVILLE
02OC016    SAINT-PIERRE (LAC) AU CURB NO. 2 <- FF

02OJ032	SAINT-LAURENT (FLEUVE) A SOREL (INCLUANT RICHELIEU)
02OJ033	SAINT-LAURENT (FLEUVE) A SOREL (SANS RICHELIEU)
02OA039    SAINT-LOUIS (LAC) A POINTE-CLAIRE

International Great Lakes Datum (1985): #106, #431


"""





import numpy as np
import datetime as dt
import pandas as pd
import sqlite3
import h5py as h5
import os



stations_niveaux = {'Jetée #1 à Montréal':'02OA046',
					'Varennes': '02OA050',
					'Sorel':'02OJ022',
					'Lac St-Pierre':'02OC016',
					#'Trois-Rivières':'',
					}

Bareas = dict(lacontario=8.4239E10, lacerie=8.6006E10, lachuron=1.91768E11, lacmichigan=1.73095E11, lacsuperior=2.10009E11, lacMHG=3.64863E11)

def marinas():
	import xlrd, re
	pat = 'Lat: (\number)\\nLong: (\number)'

	def degmin2dec(d,m,c):
		"""Convert degrees, minutes.decimals to plain decimals."""
		return int(d) + float(m + '.' + c)/60.

	with xlrd.open_workbook('../data/Suivi marinasGPSscenarios1a.xlsx') as wb:

		ws = wb.sheet_by_name('Feuil2')
		N = ws.col(0,1)
		C = ws.col(4,1)

		meta = {}
		for n, c in zip(N,C):
			if n.value == '' or c.value == '':
				continue
			d1,m1,c1,d2,m2,c2 = re.match('Lat: (-?\d+)\.(\d{2}),\s*(\d{2})\nLong: (-?\d+)\.(\d{2}),\s*(\d{2})', c.value).groups()
			lat = degmin2dec(d1,m1,c1)
			lon = -degmin2dec(d2,m2,c2)
			meta[n.value] = lon, lat
		return meta
#
def read_stations_potables():
	from openpyxl import load_workbook

	filename = '../data/AECOM/Calculs UTEP.xlsx'
	wb = load_workbook(filename=filename, data_only=True)
	ws = wb['UTEP']

	cols = ['B', 'C', 'H','I', 'U', 'V', 'AH', 'AI']
	out = {}

	for i in range(2,31):
		key = ws['A{0}'.format(i)].value
		out[key] = {}
		for c in cols:
			h = ws['{0}{1}'.format(c, 1)].value
			out[key][h] = ws['{0}{1}'.format(c, i)].value

	return out

def stations_potables():
	tmp = read_stations_potables()
	out = {}
	for key, val in tmp.items():
		out[key] = {}
		for i in [1,2,3]:
			lon, lat = val['INT{0}_LON'.format(i)], val['INT{0}_LAT'.format(i)]
			if lon and lat:
				out[key][i] = lon, lat

	return out


#
def stations_usees():
	import xlrd

	out = {}
	with xlrd.open_workbook('../data/AECOM/Calculs wwtp database check.xlsx') as wb:
		ws = wb.sheet_by_name('wwtp')
		idn = ws.col(0,1)
		ville = ws.col(1,0)

		lat = ws.col(13,0)
		lon = ws.col(14,0)


		for iv in idn:
			i = int(iv.value)
			out[i] = dict(ville=ville[i].value, lat=lat[i].value or None, lon=lon[i].value or None)

		return out
#
def stations_usees_2():
	import xlrd

	out = {}
	with xlrd.open_workbook('../data/AECOM/Calculs wwtp database_2.xlsx') as wb:
		ws = wb.sheet_by_name('STEU')
		idn = ws.col(0,0)
		ville = ws.col(2,0)

		lat = ws.col(9,0) #J
		lon = ws.col(10,0) #K


		for i in range(3,31):
			#if ws.cell_value(i,8) == 0:  continue

			out[int(idn[i].value)] = dict(nom=ville[i].value, lat=lat[i].value or None, lon=lon[i].value or None)

		return out
#
def NBS(freq='annual', stat=np.mean):
	"""Return a dictionary holding the average NBS in the reference and future
	periods for each lake and simulation.
	"""
	import ndict
	from scipy.io import matlab
	out = ndict.NestedDict()
	seasons = dict(winter=(12,1,2), spring=(3,4,5), summer=(6,7,8), fall=(9,10,11))
	months = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']

	# CRCM4
	M= matlab.loadmat('../data/NBS/CRCM4SerieNBS.mat', squeeze_me=True, struct_as_record=False)
	aliases = {}
	for per, key in zip(['ref', 'fut'], ['MSERIES_PR', 'MSERIES_FU']):
		D = M[key]
		aliases[per] = D._fieldnames

		for alias in D._fieldnames:
			d = getattr(D, alias)

			for lake in d._fieldnames:
				nbs = np.ma.masked_invalid( getattr(d, lake).nbs )
				if freq=='annual':
					out[lake][alias] = stat(nbs[:,1:].mean(1))
				elif freq in seasons.keys():
					out[lake][alias] = stat(nbs[:, seasons[freq]].mean(1))
				elif freq in months:
					out[lake][alias] = stat(nbs[:, months.index(freq)+1])


	# OBS
	#Great Lakes basin: SUP (Lake Superior), MIC (Lake Michigan), HUR (Lake Huron), G(Georgian Bay), STC (Lake St. Clair), ERI (Lake Erie), and ONT (Lake Ontario).
	O = matlab.loadmat('../data/NBS/ObsGl.mat', squeeze_me=True, struct_as_record=False)['ObsGL']
	conv = dict(lacontario='ONT', lacerie='ERI', lachuron='HUR', lacmichigan='MIC', lacsuperior='SUP', lacMHG='MHG')
	for (lake, lid) in conv.items():
		nbs = getattr(O, lid).NBSPcpLake.data
		if freq=='annual':
			out[lake]['obs'] = stat(nbs[:,-1])/365.25
		elif freq in seasons.keys():
			out[lake]['obs'] = stat(nbs[:,seasons[freq]].mean(1))/(365.25/12)
		elif freq in months:
			out[lake]['obs'] = stat(nbs[:,months.index(freq)+1])/(365.25/12)

	return out# dict(zip(aliases['ref'], aliases['fut']))

def basins_weighted_NBS_average(nbs):
	"""Return the basin weighted average."""
	import collections

	lakes = ['lacontario', 'lacMHG', 'lacerie', 'lacsuperior']
	A = sum([Bareas[l] for l in lakes])

	out = {}
	for alias in nbs.keylevel(1):
		out[alias] = sum([nbs[l][alias] * Bareas[l] for l in lakes])/A

	return out




def EC_H20(filename):
	"""Return daily time series of water levels."""
	import codecs
	from matplotlib.dates import strpdate2num, num2date
	f = codecs.open(filename)

	convert = lambda x: num2date(strpdate2num('%Y/%m/%d')(x.decode()))
	ra = np.loadtxt(f, dtype={'names':('date', 'level'), 'formats':('datetime64[D]', 'f8')}, skiprows=8, delimiter=',', usecols=[0,1], converters={0:convert, 1:float})
	return pd.TimeSeries(ra['level'], ra['date'])
